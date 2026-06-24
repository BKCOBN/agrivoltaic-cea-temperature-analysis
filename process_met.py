# matches USB0 sensor logs w/ MET station readings + writes one xlsx /month

# CONFIGURATION — edit these before running
USB0_ROOT = r"USB0"
MET_ROOT = r"German-Solar"
WINDOW_MIN = 7
MONTHS = [
    ("2025-10", "October-German-Solar", "October_MET_matched.xlsx"),
    ("2025-11", "November-German-Solar", "November_MET_matched.xlsx"),
    ("2025-12", "December-German-Solar", "December_MET_matched.xlsx"),
]

import re
from bisect import bisect_left
from datetime import datetime, timedelta, time, date
from pathlib import Path
import openpyxl

met_ts_re = re.compile(r'^(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2});(.+)$')


def read_met_csvs(folder_path):
    # read all CSVs in folder, dedupe by timestamp, return sorted (datetime, float) list
    folder = Path(folder_path)
    seen = {}
    for f in sorted(folder.glob('*.csv')):
        with open(f, 'r', encoding='utf-8-sig', errors='replace') as fh:
            for line in fh:
                line = line.rstrip('\r\n')
                m = met_ts_re.match(line)
                if not m:
                    continue
                try:
                    dt = datetime.strptime(m.group(1), '%Y/%m/%d %H:%M:%S')
                    val = float(m.group(2).replace(',', '.'))
                except ValueError:
                    continue
                if dt not in seen:
                    seen[dt] = val
    pairs = sorted(seen.items())  # list of (datetime, float), sorted
    return pairs


ts_re = re.compile(r'^(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2})(.*)', re.DOTALL)
date_re = re.compile(r'^\d{4}/\d{2}/\d{2}$')


def parse_line(line):
    # return None if line doesn't start w/ wall-clock timestamp
    m = ts_re.match(line)
    if not m:
        return None

    ldate = datetime.strptime(m.group(1), '%Y-%m-%d').date()
    ltime = datetime.strptime(m.group(2), '%H:%M:%S').time()
    rest = m.group(3).strip()

    if not rest:
        return (ldate, ltime, 'blank', None)
    if rest.startswith('Date'):
        # logger reprints headers periodically not real data
        return (ldate, ltime, 'header', None)

    parts = rest.split()
    if len(parts) < 9:
        return (ldate, ltime, 'blank', None)

    # boot msgs can slip through check date format to catch them
    if not date_re.match(parts[0]):
        return (ldate, ltime, 'blank', None)

    temps = []
    for raw in parts[3:9]:  # T01-T06 only, T07/T08 excluded (bad hardware)
        if raw.lower() == 'nan':
            temps.append(None)
        else:
            try:
                temps.append(float(raw))
            except ValueError:
                temps.append(None)
    return (ldate, ltime, 'data', temps)


def read_usb0_logs(folder_path):
    # read all USB0 TXT files in folder, return (date, time, T01..T06) readings
    folder = Path(folder_path)
    files = sorted(folder.glob('RDL_*_USB0.txt'))
    readings = []
    for f in files:
        with open(f, 'r', encoding='utf-8', errors='replace') as fh:
            for line in fh:
                parsed = parse_line(line.rstrip('\r\n'))
                if parsed is None:
                    continue
                ldate, ltime, kind, temps = parsed
                if kind != 'data':
                    continue
                readings.append((ldate, ltime, temps[0], temps[1], temps[2], temps[3], temps[4], temps[5]))
    return readings



def window_avg(keys, vals, usb0_dt, window_min):
    # mean of all MET vals w/in +-window_min mins of usb0_dt, None if no match
    lo = usb0_dt - timedelta(minutes=window_min)
    hi = usb0_dt + timedelta(minutes=window_min)
    i = bisect_left(keys, lo)
    bucket = []
    while i < len(keys) and keys[i] <= hi:
        bucket.append(vals[i])
        i += 1
    if not bucket:
        return None
    return sum(bucket) / len(bucket)


def write_workbook(rows, out_path, month_name):
    if not rows:
        raise ValueError("no rows to write")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_name
    ws.sheet_view.zoomScale = 70

    ws.append(["Date", "Time", "Soil Under Rack (°C)", "Air Under Rack (°C)", "Soil Outside (°C)", "MET Ambient (°C)", "ΔT Air-Ambient (°C)"])

    row_num = 2
    for rec_date, rec_time, soil_under, air_under, soil_outside, met_avg in rows:
        a = ws.cell(row_num, 1, datetime.combine(rec_date, time(0)))
        a.number_format = 'mm-dd-yy'
        b = ws.cell(row_num, 2, rec_time)
        b.number_format = 'h:mm:ss'
        ws.cell(row_num, 3, soil_under)
        ws.cell(row_num, 4, air_under)
        ws.cell(row_num, 5, soil_outside)
        ws.cell(row_num, 6, met_avg)
        if met_avg is not None and air_under is not None:
            ws.cell(row_num, 7, '=D%d-F%d' % (row_num, row_num))
        row_num += 1

    last_row = row_num - 1
    avg_row = last_row + 1
    for col_idx, col_letter in [(3, 'C'), (4, 'D'), (5, 'E'), (6, 'F'), (7, 'G')]:
        ws.cell(avg_row, col_idx, '=AVERAGE(%s2:%s%d)' % (col_letter, col_letter, last_row))

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                continue
            cell_len = len(cell.value) if isinstance(cell.value, str) else 10
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(out_path)


def main():
    for usb0_sub, met_sub, out_file in MONTHS:
        usb0_folder = Path(USB0_ROOT) / usb0_sub
        met_folder = Path(MET_ROOT) / met_sub

        if not usb0_folder.is_dir():
            print("skipping %s — USB0 folder not found" % usb0_sub)
            continue
        if not met_folder.is_dir():
            print("skipping %s — MET folder not found" % met_sub)
            continue

        print("Reading USB0 logs from: %s" % usb0_folder)
        usb0_readings = read_usb0_logs(usb0_folder)
        if not usb0_readings:
            print("  no USB0 readings found, skipping")
            continue

        print("Reading MET CSVs from:  %s" % met_folder)
        met_pairs = read_met_csvs(met_folder)
        if not met_pairs:
            print("  no MET readings found, skipping")
            continue

        met_keys = [p[0] for p in met_pairs]
        met_vals = [p[1] for p in met_pairs]

        rows = []
        for rec_date, rec_time, t01, t02, t03, t04, t05, t06 in usb0_readings:
            usb0_dt = datetime.combine(rec_date, rec_time)
            soil_under = (t01 + t02) / 2 if t01 is not None and t02 is not None else (t01 if t01 is not None else t02)
            air_under = (t03 + t04) / 2 if t03 is not None and t04 is not None else (t03 if t03 is not None else t04)
            soil_outside = (t05 + t06) / 2 if t05 is not None and t06 is not None else (t05 if t05 is not None else t06)
            met_avg = window_avg(met_keys, met_vals, usb0_dt, WINDOW_MIN)
            rows.append((rec_date, rec_time, soil_under, air_under, soil_outside, met_avg))

        month_name = usb0_readings[0][0].strftime("%B")
        write_workbook(rows, out_file, month_name)
        matched = sum(1 for r in rows if r[5] is not None)
        print("  wrote %s — %d rows, %d MET matches" % (out_file, len(rows), matched))


if __name__ == "__main__":
    main()
