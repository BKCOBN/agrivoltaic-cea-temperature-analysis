# CONFIGURATION — edit these two lines before running
INPUT_FOLDER = r"USB0\2025-12"
OUTPUT_FILE = r"December_generated.xlsx"

import re
from datetime import datetime, time
from pathlib import Path
import openpyxl

ts_re = re.compile(r'^(\d{4}-\d{2}-\d{2}) (\d{2}:\d{2}:\d{2})(.*)', re.DOTALL)
date_re = re.compile(r'^\d{4}/\d{2}/\d{2}$')


def parse_line(line):
    # None if line doesn't start with a wall-clock timestamp
    m = ts_re.match(line)
    if not m:
        return None

    ldate = datetime.strptime(m.group(1), '%Y-%m-%d').date()
    ltime = datetime.strptime(m.group(2), '%H:%M:%S').time()
    rest = m.group(3).strip()

    if not rest:
        return (ldate, ltime, 'blank', None)
    if rest.startswith('Date'):
        # logger reprints headers periodically, not real data
        return (ldate, ltime, 'header', None)

    parts = rest.split()
    if len(parts) < 9:
        return (ldate, ltime, 'blank', None)

    # boot messages can slip through, catch them by checking the date format
    if not date_re.match(parts[0]):
        return (ldate, ltime, 'blank', None)

    temps = []
    for raw in parts[3:9]:  # T01-T06 only, T07/T08 excluded (bad hardware)
        if raw.lower() == 'nan':
            temps.append(None)
        else:
            try:
                temps.append(float(raw))
            except ValueError:
                temps.append(None)
    return (ldate, ltime, 'data', temps)


def read_logs(folder_path):
    # hdr = timestamp of the first header line seen
    folder = Path(folder_path)
    files = list(folder.glob('RDL_*_USB0.txt'))
    files.sort()  # alphabetical = chronological (YYYY-MM-DD names)
    hdr = None
    readings = []

    for f in files:
        with open(f, 'r', encoding='utf-8', errors='replace') as fh:
            for line in fh:
                parsed = parse_line(line.rstrip('\r\n'))
                if parsed is None:
                    continue
                ldate, ltime, kind, temps = parsed
                if kind == 'header' and hdr is None:
                    hdr = (ldate, ltime)
                elif kind == 'data':
                    readings.append((ldate, ltime) + tuple(temps))
    return hdr, readings


def write_workbook(hdr, readings, out_path, month):
    if not readings:
        raise ValueError("no readings")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month
    ws.sheet_view.zoomScale = 70

    # row 1: timestamp from first header + sensor/group labels
    ts_date, ts_time = hdr
    ws['A1'] = datetime.combine(ts_date, time(0))
    ws['A1'].number_format = 'mm-dd-yy'
    ws['B1'] = ts_time
    ws['B1'].number_format = 'h:mm:ss'
    labels = ['T01', 'T02', 'Under Modules inside Pots', 'T03', 'T04', 'Under Modules', 'T05', 'T06', 'Control']
    col_num = 3
    for label in labels:
        ws.cell(row=1, column=col_num, value=label)
        col_num += 1

    # data rows, E/H/K get AVERAGE formulas over the paired raw sensor cols
    row_num = 2
    for reading in readings:
        rec_date, rec_time, t01, t02, t03, t04, t05, t06 = reading
        a = ws.cell(row_num, 1, datetime.combine(rec_date, time(0)))
        a.number_format = 'mm-dd-yy'
        b = ws.cell(row_num, 2, rec_time)
        b.number_format = 'h:mm:ss'
        ws.cell(row_num, 3, t01)
        ws.cell(row_num, 4, t02)
        ws.cell(row_num, 5, '=AVERAGE(C%d:D%d)' % (row_num, row_num))
        ws.cell(row_num, 6, t03)
        ws.cell(row_num, 7, t04)
        ws.cell(row_num, 8, '=AVERAGE(F%d:G%d)' % (row_num, row_num))
        ws.cell(row_num, 9, t05)
        ws.cell(row_num, 10, t06)
        ws.cell(row_num, 11, '=AVERAGE(I%d:J%d)' % (row_num, row_num))
        row_num += 1

    # last row: overall avg per column
    last_row = row_num - 1
    avg_row = last_row + 1
    avg_col_map = [
        (3, 'C'), (4, 'D'), (5, 'E'),
        (6, 'F'), (7, 'G'), (8, 'H'),
        (9, 'I'), (10, 'J'), (11, 'K'),
    ]
    for col_idx, col_letter in avg_col_map:
        ws.cell(avg_row, col_idx, '=AVERAGE(%s2:%s%d)' % (col_letter, col_letter, last_row))

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value == None:
                continue
            if isinstance(cell.value, str) and cell.value.startswith('='):
                continue
            if isinstance(cell.value, str):
                cell_len = len(cell.value)
            else:
                cell_len = 10
            if cell_len > max_len:
                max_len = cell_len
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(out_path)


def main():
    folder = Path(INPUT_FOLDER)
    if not folder.is_dir():
        raise SystemExit("can't find folder: %s" % INPUT_FOLDER)

    print("Reading TXT files from: %s" % folder.resolve())
    hdr, readings = read_logs(folder)
    if not readings:
        raise SystemExit("no readings found, check INPUT_FOLDER")
    if hdr == None:
        raise SystemExit("no header line found in any txt file")
    month = readings[0][0].strftime("%B")
    write_workbook(hdr, readings, OUTPUT_FILE, month)
    print("Done.")

if __name__ == "__main__":
    main()
