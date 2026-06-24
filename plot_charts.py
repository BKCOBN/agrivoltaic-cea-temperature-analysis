from datetime import datetime, date, time
from pathlib import Path
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference, Series

# reads *_MET_matched.xlsx +- writes Summary_Charts.xlsx w/ 7 sheets

# update these paths if your filenames differ
MONTHS = [
    ("October_MET_matched.xlsx", "October"),
    ("November_MET_matched.xlsx", "November"),
    ("December_MET_matched.xlsx", "December"),
]
OUT_FILE = "Summary_Charts.xlsx"

LABELS = ["Soil Under Rack", "Air Under Rack", "Soil Outside", "MET Ambient"]
COLORS = ["4472C4", "ED7D31", "70AD47", "FF0000"]  # blue, orange, green, red

DAY_TABS = {
    "October": "Oct Day",
    "November": "Nov Day",
    "December": "Dec Day",
}

DT_TABS = {
    "October": "Oct High dT",
    "November": "Nov High dT",
    "December": "Dec High dT",
}


def avg(vals):
    # skip None vals
    total = 0.0
    count = 0
    for v in vals:
        if v is not None:
            total += v
            count += 1
    if count == 0: return 0.0
    return total / count


def read_xlsx(path):
    # read *_MET_matched.xlsx, cols C=soil_under, D=air_under, E=soil_outside, F=met_ambient
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 6 or row[0] is None:
            continue
        dval = row[0]
        rdate = None
        if isinstance(dval, datetime):
            rdate = dval.date()
        elif isinstance(dval, date):
            rdate = dval
        if rdate is None:
            continue
        tv = row[1]
        rtime = None
        if isinstance(tv, time):
            rtime = tv
        elif isinstance(tv, datetime):
            rtime = tv.time()
        elif isinstance(tv, float):
            sec = round(tv * 86400)
            rtime = time(sec // 3600 % 24, sec % 3600 // 60, sec % 60)
        if rtime is None:
            continue
        rows.append((rdate, rtime, row[2], row[3], row[4], row[5]))
    return rows


def rep_day(data):
    # pick day whose mean MET ambient is closest to monthly mean MET ambient
    if not data:
        return None
    monthly_met = avg([r[5] for r in data])

    by_day = {}
    for r in data:
        if r[0] not in by_day:
            by_day[r[0]] = []
        by_day[r[0]].append(r)

    rdate = None
    mdist = None
    for d, day in by_day.items():
        daily_met = avg([r[5] for r in day])
        dist = abs(daily_met - monthly_met)
        if mdist is None or dist < mdist:
            mdist = dist
            rdate = d
    return rdate


def peak_dt_day(data):
    # find day where under-rack air ran hottest relative to MET ambient
    if not data:
        return None
    by_day = {}
    for r in data:
        if r[0] not in by_day:
            by_day[r[0]] = []
        by_day[r[0]].append(r)
    pdate = None
    mdt = None
    for d, day in by_day.items():
        dts = [r[3] - r[5] for r in day if r[3] is not None and r[5] is not None]
        dt = avg(dts)
        if mdt is None or dt > mdt:
            mdt = dt
            pdate = d
    return pdate


def day_sheet(wb, name, day_rows):
    # format time as string, Excel treats raw time vals as numbers
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    ws.append(["Time"] + LABELS)
    for _, t, su, au, so, met in day_rows:
        ws.append([t.strftime("%H:%M:%S"), su, au, so, met])
    return ws


def bar_chart(ws, n):
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Monthly Average Temperatures by Treatment"
    chart.y_axis.title = "Temperature (°C)"
    chart.x_axis.title = "Month"
    chart.width = 22
    chart.height = 14
    col_idx = 2
    for color in COLORS:  # one series /treatment group
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=n + 1)
        series = Series(data, title_from_data=True)
        series.graphicalProperties.solidFill = color
        chart.series.append(series)
        col_idx += 1
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n + 1))
    return chart


def line_chart(ws, title):
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Temperature (°C)"
    chart.x_axis.title = "Time of Day"
    chart.width = 22
    chart.height = 14
    n = ws.max_row
    col_idx = 2
    for color in COLORS:
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=n)
        series = Series(data, title_from_data=True)
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.line.width = 20000
        chart.series.append(series)
        col_idx += 1
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=n))
    return chart


def write_workbook(months, out):
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["Month"] + LABELS)
    for m in months:
        ws_sum.append([m["name"], m["means"][0], m["means"][1], m["means"][2], m["means"][3]])
    ws_sum.add_chart(bar_chart(ws_sum, len(months)), "F2")

    for m in months:
        tab = DAY_TABS.get(m["name"], "%s Day" % m["name"])
        ws = day_sheet(wb, tab, m["rep_rows"])
        ws.add_chart(line_chart(ws, "%s - Representative Day (%s)" % (m["name"], m["rep_date"])), "F2")
    for m in months:
        tab = DT_TABS.get(m["name"], "%s High dT" % m["name"])
        ws = day_sheet(wb, tab, m["dt_rows"])
        ws.add_chart(line_chart(ws, "%s - Highest ΔT Day (%s)" % (m["name"], m["peak"])), "F2")
    wb.save(out)


def main():
    months = []
    for fpath, mname in MONTHS:
        p = Path(fpath)
        if not p.exists():
            print("%s not found, skipping" % p)
            continue
        print("Reading %s ..." % p)
        data = read_xlsx(fpath)
        if not data:
            print("no data in %s, skipping" % p)
            continue

        rdate = rep_day(data)
        peak = peak_dt_day(data)

        # grab means + rep day rows + peak day rows in one pass
        rep_rows = [r for r in data if r[0] == rdate]
        dt_rows = [r for r in data if r[0] == peak]

        months.append({
            "name": mname,
            "means": (avg([r[2] for r in data]), avg([r[3] for r in data]),
                      avg([r[4] for r in data]), avg([r[5] for r in data])),
            "rep_date": rdate,
            "rep_rows": rep_rows,
            "peak": peak,
            "dt_rows": dt_rows,
        })
        print("  -> rep day: %s (%d readings)" % (rdate, len(rep_rows)))
        print("  -> peak dT: %s (%d readings)" % (peak, len(dt_rows)))

    if not months:
        raise SystemExit("nothing to process, check your xlsx files")
    write_workbook(months, OUT_FILE)
    print("Done.")

if __name__ == "__main__":
    main()
